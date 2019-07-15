"""
    Author:         <REPLACE>
    Project:        100DaysPython
    File:           module3_day32_excel.py
    Creation Date:  <REPLACE>
    Description:    <REPLACE>
"""

import openpyxl

sales_wb = openpyxl.load_workbook('Sales.xlsx')
customers_wb = openpyxl.load_workbook('CustomerDIM.xlsx')
products_wb = openpyxl.load_workbook('ProductDIM.xlsx')

sales = sales_wb["sales"]
customers = customers_wb["customer_dim"]
products = products_wb["product_dim"]

print(customers["C5"].value)

sales["E1"] = "first_name"
sales["F1"].value = "last_name"
sales["G1"] = "product_name"
sales["H1"] = "product_price"
sales["I1"] = "subtotal"
sales["J1"] = "tax"
sales["K1"] = "total"

for i, sr in enumerate(sales.rows, start=2):
    for j, cr in enumerate(customers.rows, start=2):
        if sales[f"B{i}"].value == cr[0].value:
            sales[f"E{i}"] = cr[1].value
            sales[f"F{i}"] = cr[2].value

    for k, pr in enumerate(products.rows, start=2):
        if sales[f"C{i}"].value == pr[0].value:
            sales[f"G{i}"] = pr[1].value
            sales[f"H{i}"] = pr[2].value

    try:
        sales[f"I{i}"] = sales[f"H{i}"].value * sales[f"D{i}"].value
        sales.cell(row=i, column=9).number_format = "$#,##0.00"

        sales[f"J{i}"] = sales[f"I{i}"].value * 0.06
        sales.cell(row=i, column=10).number_format = "$#,##0.00"

        sales[f"K{i}"] = sales[f"I{i}"].value + sales[f"J{i}"].value
        sales.cell(row=i, column=11).number_format = "$#,##0.00"
    except TypeError:
        pass

sales_wb.save("Sales.xlsx")
